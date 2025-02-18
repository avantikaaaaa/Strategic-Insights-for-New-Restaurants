import pandas as pd
import os
import re
import openpyxl
import time
from openpyxl import Workbook

start_time = time.time()
# Delete the previously created file
if os.path.exists("JobFilesList2.xlsx"):
    os.remove("JobFilesList2.xlsx")
#
pathXX = os.path.abspath(os.getcwd())
pathdir = pathXX
#
# Create an excel workbook
wb = Workbook()
ws = wb.active

# open the input excel workbook
input_wb = openpyxl.load_workbook("InputJobist.xlsx")

# Referring the sheet within the workbook
input_sheet = input_wb['Sheet1']

# Get the max row and column value from the Input sheet
mrow = input_sheet.max_row
mcol = input_sheet.max_column
num_of_row_written = 0

get_overrides_from_block = False
block_lines = ''

global records, job_location

# df_output = pd.DataFrame()
records = []
job_location = ''


###***************************************************
### Write output file
###***************************************************
#
def write_output_file(application, schedule, lob, JOB_NAME, PROC_NAME, FILE_TYPE, STEP_NAME, FILE_NAME, LOGICAL_FILE,)
                       utility, JOB_COMMENT, jcl_step_name, disposition):
    global isn, parent_folder, data_center

    # if FILE_TYPE == 'Input File':
    #     FILE_TYPE = 'Read'
    #
    # if FILE_TYPE == 'Output File':
    #     FILE_TYPE = 'Write'
    global records
    num_of_row_written = 1

    records = {
        'ISN': isn,
        'application': application,
        'schedule': schedule,
        'parent folder': parent_folder,
        'data center': data_center,
        'job': JOB_NAME,
        'job step': jcl_step_name,
        'proc': PROC_NAME,
        'step': STEP_NAME,
        'program': utility,
        'program-type': '',
        'secondary program': '',
        'secondary type': '',
        'resource': FILE_NAME,
        'resource-type': 'FILE',
        'access-type': FILE_TYPE,
        'LOB': lob,
        'Logical file name': LOGICAL_FILE,
        'Comeent': JOB_COMMENT,
        'File Disposition': disposition
        # ADD more variables/columns as needed
      }

      records.append(record)
      return num_of_row_written


def extract_lines_between_exec(file_path, JOB_NAME):
    exec_blocks = []
    current_block = []

    try:
        with open(file_path, encoding='utf-8', errors='ignore') as file:
            for line in files:
                if 'EXEC' in line:
                    if current_block:
                        exec_blocks.append(current_block)
                        current_block = []
                    current_block.append(line.strip())
                elif current_block:
                    if line[0:3] == "// ":
                        # print(line)
                        characters = line[2:70] # Extract characters from the 3rd to 72nd position
                        if characters.isspace():
                            # lines_missed = len(lines) -l1
                            # print(lines_missed)
                            # if lines_missed > 4;
                            # JOB_COMMENT = '// Null line found in JCL/PROC'
                            # write_output_file(application, schedule, lob, JOB_NAME, PROC_NAME, '', '', '', '',
                            #                    '',JOB_COMMENT, '', '')
                            # JOB_COMMENT = ''
                            break
                        current_block.append(lines.strip())

        # Add the last block if it exists
        if current_block:
            exec_blocks.append(current_block)

        return exec_blocks
    except FileNotFoundError:
        JOB_COMMENT = 'JOB Not found'
        rows_written = write_output_file(application, schedule, lob, JOB_NAME, '','','','','','', JOB_COMMENT, ''
                                         '')

        JOB_COMMENT = ''
        return exec_blocks


def extract_jcl_steps(path, JOB_NAME):
    PROC_LIST = []
    PROC_NAME = ''
    LOGICAL_FILE = ''
    proc_step_name = ''
    global get_overrides_from_block, block_lines, application, schedule, lob, num_of_row_written

    srch_txt_exec = ' EXEC '
    JOB_COMMENT = ''
    # print('srch jcl name : ', job_name)
    file_path = path

    # Extract the JCL steps (Lines b/w each ' EXEC ' keyword) as blocks
    blocks = extract_lines_between_exec(file_path, JOB_NAME)
    get_overrides_from_block = True

    # Extract steps and Files from JCL 1st, if the step is not executing the PROC
    check_jcl = True
    INPUT_FILE_LIST, INPUT_STEP_LIST, LOGICAL_FILE_LIST, ACTUAL_UTILITY_LIST, disposition_list, access_type_list = extract_files(
        JOB_NAME, '', check_jcl, proc_step_name)
    # if len(INPUT_FILE_LIST) > 0:
    #    print('MAIN Job name : ', MAIN_JOBNAME)
    #    print('MAIN PROC List : ', MAIN_PROC_NAME)
    #    print('Input files list : ', INPUT_FILE_LIST)
    #    print('Input step list : ', INPUT_STEP_LIST)
    #    print('LOGICAL_FILE_LIST : ', LOGICAL_FILE_LIST)

    if len(INPUT_FILE_LIST) > 0:
        J = 0
        for input_file in INPUT_FILE_LIST:
            num_of_row_written = write_output_file(application, schedule, lob, JOB_NAME, '',access_type_list[j],
                                        '',INPUT_FILE_LIST[j],LOGICAL_FILE_LIST[j],
                                                   ACTUAL_UTILITY_LIST[j], JOB_COMMENT, INPUT_STEP_LIST[j]
                                                   disposition_list[j])
            j = j + 1

    # iterate through each extracted JCL steps and find for PROCs
    # if len(blocks) > 0:
    for block in blocks:
        block_lines = block
        PROC_LIST = []
        proc_step_list = []
        proc_step_name = ''

        # print('get_overrides_from_block: ', get_overrides_from_block)
        # print("---- Block Start ----")
        for lines in blocks:
            # print(line)
            # check if string present on a current line
            if line.find(srch_txt_exec) != -1:
                #                print(srch_txt_exec, 'string exists in file')
                #                print('Line Number:', lines.index(line))
                #                print('Line:', line)
                if line[0:3] != "//*":
                    split_line = line.split()
                    if len(split_line) > 2:
                        PROC_NAME = split_line[2]
                        proc_step_name = split_line[0]
                    #                        print(PROC_NAME)
                    #                    print(PROC_NAME)
                    if PROC_NAME.find(',') != -1:
                        PROC_NAME = PROC_NAME.split(',')
                        PROC_NAME = PROC_NAME[0]

                    if PROC_NAME.find('=') != -1:
                        PROC_NAME =PROC_NAME[PROC_NAME.find("=") + 1:]

                    PROC_NAME =PROC_NAME[:8]
                    proc_step_name = proc_step_name[2:]
                    proc_step_name = proc_step_name.split('')[0]

                    if PROC_NAME is not None:
                        if PROC_NAME.find(':') != -1:
                            pass
                        else:
                            if len(PROC_NAME) > 4:
                                if PROC_NAME not in ['IDCAM', 'IEFBR14', 'IEBGENER', 'SORT', 'FLUSH',
                                                     'SAS', 'EZTPA00', 'IKJEFT01', 'IDCAMS2', 'CMPRORG',
                                                     'IEBCOPY', 'NDMBATCH', 'INFOWAIT', 'SQL', 'EXECUTE',
                                                     'ICETOOL', 'JZOSPROC', 'EXECUTIN', 'FTP', 'IKJEFT1B',
                                                     'EZACFSM1', 'MISPSMTP', 'DB2CPY1P'] and PROC_NAME[3] == 'P':
                                    PROC_LIST.append(PROC_NAME)
                                    proc_step_list.append(proc_step_name)
        # print("----Block End ----")
        # print('PROC_LIST: ', PROC_LIST)
        MAIN_JOBNAME = JOB_NAME

        # Extract the files and steps from the PROCs identified
        if len(PROC_LIST) == 1:
            MAIN_PROC_NAME = PROC_LIST[0]
            check_jcl = False
            INPUT_FILE_LIST, INPUT_STEP_LIST, LOGICAL_FILE_LIST, ACTUAL_UTILITY_LIST, disposition_list, access_type_list = extract_files(MAIN_JOBNAME, MAIN_PROC_NAME, check_jcl, proc_step_name)
            # if len(INPUT_FILE_LIST) > 0:
            #    print('MAIN job name :', MAIN_JOBNAME)
            #    print('Main PROC List :', MAIN_PROC_NAME)
            #    print('Input files list :', INPUT_FILE_LIST)
            #    print('Input Step list :', INPUT_STEP_LIST)
            #    print('Logical_file_list :', LOGICAL_FILE_LIST)

            if len(INPUT_FILE_LIST) > 0:
                I = 0
                for input_file in INPUT_FILE_LIST:
                    num_of_row_written = write_output_file(application, schedule, lob, MAIN_JOBNAME, MAIN_PROC_NAME, access_type_list[i],
                                                           INPUT_STEP_LIST, INPUT_FILE_LIST, LOGICAL_FILE_LIST[i], ACTUAL_UTILITY_LIST[i], JOB_COMMENT, proc_step_name, disposition_list[i])
                    i = i + 1

def check_jcl_full_file_override(JOB_NAME, PROC_NAME, ACTUAL_INPUT_FILE_LIST, INPUT_STEP_LIST, logical_file_name_list, FORMATTED_FILE_LIST, qualifier_override):
    jclpath = r"C:\Users\UX169817\PyCharm Project\Automation\tar\20231205_vg-project-all-code"
    os.chdir(jclpath)
    global job_location
    tmp_jobname = JOB_NAME
    job_name = tmp_jobname.upper() + '.JCL'
    JOB_NAME = tmp_jobname.upper()
    PROC_NAME = []
    global foundinJCL
    foundinJCL = False
    global MODULE_NAME, get_overrides_from_block, block_lines
    MODULE_NAME = ''
    module_name_temp = ''
    srch_dddsn = 'DSN='
    proc_match = False
    prev_logical_file_name = ''

    # print('inside check_full_file_override')
    if get_overrides_from_block is True:
        lines = block_lines   fp.readlines()

    else:
        with open(f"{jclpath}\{job_name}", encoding='utf-8', errors='ignore') as fp:
            lines = fp.readlines()

    for line in lines:
        # print('Lines :' line)
        if line[0:3] != "//*":
            if line.find(' EXEC ') != -1:

                if line.find(PROC_NAME) != -1:
                    # print("jcl override line: ", line)
                    proc_match = True
                else:
                    proc_match = False
            if line.find(srch_dddsn) != -1 and proc_match is True and 'PARM' not in line:
                if quaifier_override is True and '&' not in line:
                    file_name = line[line.find("DSN=") + 4:line.find(",")]
                    step_name = line[line.find("//") + 2:line.find(" DD")]
                    # print("JCL Override file: ", file_name)
                    # print("JCL Override step: ", step_name)
                    if ' ' in file_name:
                        file_name = file_name.split(' ')[0]

                    if '(' in file_name:
                        file_name = file_name[:file_name.find("(")]
                    split_step_file = step_name.split('.')
                    # print(len(split_step_file))
                    # print(split_step_file)
                    if len(split_step_file) > 1:
                        logical_file_name = split_step_file[1]
                        logical_file_name = logical_file_name.strip()
                        if logical_file_name == '':
                            logical_file_name = prev_logical_file_name

                        else:
                            prev_logical_file_name = logical_file_name
                        # print(logical_file_name)
                        if logical_file_name in logical_file_name_list:
                            index_num = logical_file_name_list.index(logical_file_name)
                            ACTUAL_INPUT_FILE_LIST[index_num] = file_name

                elif qualifier_override is False and '&' in line:
                    # file_name = line[line.find("DSN=") + 4:line.find(",")]
                    file_name = line[line.find("DSN=") + 4:]
                    if ',' in file_name:
                        split_file = file_name.split(',')
                        file_name = split_file[0]

                    step_name = line[line.find("//") + 2:line.find(" DD")]
                    # print("JCL Override file: ", file_name)
                    # print("JCL Override step: ", step_name)
                    if ' ' in file_name:
                        file_name = file_name.split(' ')[0]

                    if '(' in file_name:
                        file_name = file_name[:file_name.find("(")]
                    split_step_file = step_name.split('.')
                    # print(len(split_step_file))
                    # print(split_step_file)
                    if len(split_step_file) > 1:
                        logical_file_name = split_step_file[1]
                        logical_file_name = logical_file_name.strip()
                        if logical_file_name == '':
                            logical_file_name = prev_logical_file_name

                        else:
                            prev_logical_file_name = logical_file_name
                        # print(logical_file_name)
                        if logical_file_name in logical_file_name_list:
                            index_num = logical_file_name_list.index(logical_file_name)
                            # print(logical_file_name)
                            # print('INPUT_FILE_LIST', ACTUAL_INPUT_FILE_LIST)
                            formatted_filr_name = handle_aphrnd_period(file_name)
                            ACTUAL_INPUT_FILE_LIST[index_num] = file_name
                            FORMATTED_FILE_LIST[index_num] = formatted_filr_name


    return ACTUAL_INPUT_FILE_LIST, FORMATTED_FILE_LIST


def check_set_from_jcl(temp_file, JOB_NAME):
    jcl_path = r"C:\Users\UX169817\PyCharm Project\Automation\tar\20231205_vg-project-all-code"
    global job_location
    tmp_jobname = JOB_NAME
    tmp_jobname = re.sub('\w+', '', tmp_jobname)
    tmp_jobname = tmp_jobname.upper() + '.JCL'

    word1 = temp_file[1:]
    MODULE_NAME = temp_file

    with open(f"{jclpath}\{job_name}", encoding='utf-8', errors='ignore') as fp
        jcl_content = file.read()
        # print('Read jcl_content')

    # print('word1 : ', word1)
    if ' SET ' in jcl_content:
        with open(f"{jclpath}\{job_name}", encoding='utf-8', errors='ignore') as fp
        lines = fp.readlines()
        for line in lines:

            if ' SET ' in line:

                pattern = rf"{word}=(.*)"

                match = re.search(pattern, line)
                if match:
                    words = match.group(1)
                    # print('words: ', words)
                    MODULE_NAME =words

                    if '\n' in MODULE_NAME:
                        MODULE_NAME = MODULE_NAME[:-1]

                    if ' ' in MODULE_NAME:
                        split_second_string = MODULE_NAME.split()
                        MODULE_NAME = split_second_string[0]

    return MODULE_NAME


def check_jcl_override(JOB_NAME, temp_module_name, PROC_NAME):
    jcl_path = r"C:\Users\UX169817\PyCharm Project\Automation\tar\20231205_vg-project-all-code"
    global job_location
    os.chdir(jclpath)
    tmp_jobname = JOB_NAME
    tmp_jobname = re.sub('\w+', '', tmp_jobname)
    job_name = tmp_jobname.upper() + '.JCL'
    #
    JOB_NAME = tmp_jobname.upper()
    PROC_NAME = []
    global foundinJCL
    foundinJCL = False
    global MODULE_NAME, get_overrides_from_block, block_lines
    MODULE_NAME = ''
    module_name_temp = ''
    proc_match = False
    jcl_content = ''

    # print('inside check_override')
    # print('temp_module_name: ', temp_module_name)
    # print('get_overrides_from_block: ', get_overrides_from_block)

    if '\n' in temp_module_name:
        temp_module_name = temp_module_name[:-1]

    if get_overrides_from_block is True:
        lines = block_lines

    else:
        with open(f"{jclpath}\{job_name}", encoding='utf-8', errors='ignore') as fp
            lines = fp.readlines()

    with open(f"{jclpath}\{job_name}", encoding='utf-8', errors='ignore') as file:
            jcl_content = file.read()
            # print('Read jcl_content')

    if ' SET ' in jcl_content:
       # print('Read jcl_content')

        for line in lines:
            word1 = temp_module_name

            pattern = rf"{word}=(.*)"

            match = re.search(pattern, line)
            if match and line[line.find(temp_module_name)-1].isalnum() == False:
                words = match.group(1)
                # print('words: ', words)
                MODULE_NAME= words

            if '\n' in MODULE_NAME:
                MODULE_NAME = MODULE_NAME[:-1]

            if ',' in MODULE_NAME:
                MODULE_NAME = MODULE_NAME.split(',')[0]

            if ' ' in MODULE_NAME:
                split_second_string = MODULE_NAME.split()
                MODULE_NAME = split_second_string[0]
                if ',' in MODULE_NAME:
                    MODULE_NAME = MODULE_NAME[:-1]
                    if ',' in MODULE_NAME:
                        MODULE_NAME = MODULE_NAME[:-1]
                # print('Found Space.. Breaking', MODULE_NAME)
                PARMfoundinJCL = True
                foundinJCL = True
                break

            if ',' in MODULE_NAME:
                # print('Checking ","')
                MODULE_NAME = MODULE_NAME[:-1]
                if ',' in MODULE_NAME:
                    MODULE_NAME = MODULE_NAME[:-1]
                    if ',' in MODULE_NAME:
                        MODULE_NAME = MODULE_NAME[:-1]
            foundinJCL = True
            break
    else:
        # print("check SET in else part")
        if get_overrides_from_block is True:
            with open(f"{jclpath}\{job_name}", encoding='utf-8', errors='ignore') as fp:
                lines = fp.readlines()

        for line in lines:
            # print('Lines :' line)
            if line[0:3] != "//*":
                # print('temp_module_name', temp_module_name)
                if line.find(' EXEC ') != -1:

                    if line.find(PROC_NAME) != -1:
                        # print("PROC match: ", line)
                        proc_match = True
                    else:
                        proc_match = False

                find_qualifier = temp_module_name + '='
                # print('find_qualifier: ', find_qualifier)
                if line.find(find_qualifier) != -1 and line.find(' SET ') != -1:
                    split_lineproc = line.split(',')
                    # print('lineproc', line)
                    # print(len(split_lineproc), split_lineproc)
                    if len(split_lineproc) > 0:
                        for each_split in split_lineproc:
                            # print('each_split ', each_split)
                            if each_split.find(find_qualifier) != -1 and each_split[each_split.find(find_qualifier)-1].isalnum() == False:
                               if each_split.find("'") != -1:
                                  # print("found ''' ", each_split)
                                  # print(each_split.count("'"))
                                  if each_split.find("'") > 1:
                                     # MODULE_NAME = each_split[each_split.find("=") + 2:each_split.find(",") - 1]
                                     match = re.search(r"'(.*?)'", each_split)
                                     MODULE_NAME = match.group(1) input_sheet match else ''
                                     # print(MODULE_NAME)
                                     PARMfoundinJCL = True
                                     foundinJCL = True
                                     break

                               else:
                                   match = re.search(r'(.+)=([^=]*)', each_split)

                                   if match and each_split[each_split.find(find_qualifier)-1].isalnum() == False:
                                       # extract the second string part
                                       MODULE_NAME = match.group(2)
                                       foundinJCL = True
                                       # print the second string part
                                       # print(MODULE_NAME)
                                       if '\n' in MODULE_NAME:
                                           MODULE_NAME = MODULE_NAME[:-1]

                                       if ' ' in MODULE_NAME:
                                           split_second_string = MODULE_NAME.split()
                                           MODULE_NAME = split_second_string[0]
                                           if ',' in MODULE_NAME:
                                               MODULE_NAME = MODULE_NAME[:-1]
                                               if ',' in MODULE_NAME:
                                                   MODULE_NAME = MODULE_NAME[:-1]
                                           # print('Found Space.. Breaking', MODULE_NAME)
                                           PARMfoundinJCL = True
                                           foundinJCL = True
                                           break

                                       if ',' in MODULE_NAME:
                                           # print('Checking ","')
                                           MODULE_NAME = MODULE_NAME[:-1]
                                           if ',' in MODULE_NAME:
                                               MODULE_NAME = MODULE_NAME[:-1]
                                               if ',' in MODULE_NAME:
                                                   MODULE_NAME = MODULE_NAME[:-1]
                                           foundinJCL = True
                                           break

    if foundinJCL is False:
        if get_overrides_from_block is True:
            lines = block_lines

        else:
            with open(f"{jclpath}\{job_name}", encoding='utf-8', errors='ignore') as fp:
                lines = fp.readlines()

            for line in lines:
                # print('Not found in SET')
                # print('Lines :' , line)
                if line[0:3] != "//*":
                    # print('temp_module_name', temp_module_name)
                    if line.find(' EXEC ') != -1:

                        if line.find(PROC_NAME) != -1:
                            # print("PROC match: ", line)
                            proc_match = True
                        else:
                            proc_match = False

                    find_qualifier = temp_module_name + '='
                    # print('find_qualifier: ',find_qualifier)
                    if line.find(find_qualifier) != -1 and proc_match is True:
                        split_lineproc = line.split(',')
                        # print('lineproc', line)
                        # print(len(split_lineproc), split_lineproc)
                        if len(split_lineproc) > 0:
                            for each_split in split_lineproc:
                                # print('each_split: ', each_split)
                                if each_split.find(find_qualifier) != -1 and each_split[each_split.find(find_qualifier)-1].isalnum() == False:
                                   if each_split.find("'") != -1:
                                       # print("found ''' ". each_split)
                                       # print(each_split.count("'"))
                                       if each_split.count("'") > 1:
                                          # MODULE_NAME = each_split[each_split.find("=") + 2:each_split.find(",") - 1]
                                          match = re.search(r"'(.*?)'", each_split)
                                          MODULE_NAME = match.group(1) if match else ''
                                          # print(MODULE_NAME)
                                          PARMfoundinJCL =True
                                          foundinJCL =True
                                          break

                                   else:
                                       # print('inside else')
                                       match = re.search(r'(.+)=([^=]*)', each_split)

                                       if match and each_split[each_split.find(find_qualifier) - 1].isalnum() == False:
                                           # extract the second string part
                                           MODULE_NAME = match.group(2)
                                           foundinJCL = True
                                           # print the second string part
                                           # print(MODULE_NAME)
                                           if '\n' in MODULE_NAME:
                                               MODULE_NAME = MODULE_NAME[:-1]

                                           if ' ' in MODULE_NAME:
                                               split_second_string = MODULE_NAME.split()
                                               MODULE_NAME = split_second_string[0]
                                               if ',' in MODULE_NAME:
                                                   MODULE_NAME = MODULE_NAME[:-1]
                                                   if ',' in MODULE_NAME:
                                                       MODULE_NAME = MODULE_NAME[:-1]
                                                       # print('Found Space.. Breaking', MODULE_NAME)
                                                       PARMfoundinJCL = True
                                                       foundinJCL = True
                                                       break

                                                   if ',' in MODULE_NAME:
                                                       # print('Checking ","')
                                                       MODULE_NAME = MODULE_NAME[:-1]
                                                       if ',' in MODULE_NAME:
                                                           MODULE_NAME = MODULE_NAME[:-1]
                                                           if ',' in MODULE_NAME:
                                                               MODULE_NAME = MODULE_NAME[:-1]
                                                       foundinJCL = True
                                                       break
                                               # MODULE_NAME = each_split[each_split.find("=") + 1:each_split.find(",")]
                                               # print(MODULE_NAME)
                                               # if len(MODULE_NAME) > 8:
                                               #     module_name_temp =MODULE_NAME.split()
                                               #     MODULE_NAME = module_name_temp[0]
                                               #
                                               # MODULE_NAME =re.sub('\W+', '', MODULE_NAME)
                                               # MODULE_NAME = MODULE_NAME[0:8]
                                               # PARMfoundinJCL =True
                                               # foundinJCL = True
                                               # break

    # print('JCL override', MODULE_NAME, foundinJCL)
    if '&' in MODULE_NAME:
        foundinJCL = False

    # print('JCL override', MODULE_NAME, foundinJCL)
    return MODULE_NAME, foundinJCL


def check_proc_name(tmp_jobname):
    jcl_path = r"C:\Users\UX169817\PyCharm Project\Automation\tar\20231205_vg-project-all-code"
    os.chdir(jclpath)
    # print('temp_jobname: ', tmp_jobname)
    tmp_jobname = re.sub('\w+', '', tmp_jobname)


    job_name = tmp_jobname.upper() + '.JCL'
    # PROC_LIST = []
    global get_overrides_from_block, check_jcl_flag, application, schedule, lob, job_location
    get_overrides_from_block = False
    check_jcl_flag = True


    # print('srch jcl name : ',job_name)

    extract_jcl_steps(f"{jclpath}\{job_name}", JOB_NAME)


#
# Get Qualiders to be replaced
#
def get_qualifiers(INPUT_FILE_LIST):

    qualifier_list = []

    if INPUT_FILE_LIST is not None and len(INPUT_FILE_LIST) > 0:
        for file1 in INPUT_FILE_LIST:
            split_filename = file1.split('.')
            split_filename = split_filename
            # print(split_filename)
            for qualifiers in split_filename:
                if qualifiers.find('&') != -1:
                    if qualifiers[0:2] == '&&':
                        pass
                        # print('Temp file', qualifiers)
                    else:
                        # print('qualifiers: ', qualifiers)
                        qualifier_list.append(qualifiers)

        qualifier_list = list(set(qualifier_list))
        # print('qualifier_list: ', qualifier_list)
        for i in range(len(qualifier_list)):
            # print(qualifier_list[i])
            #  if qualifier_list[i][0:2] == '&&':
            #      print('Temp file')
            count = qualifier_list[i].count('&')
            if count > 1:
                new_items = qualifier_list[i].split('&')
                # print('new_items: ', new_items)
                new_items = [item1 for item1 in new_items if item1 != '']
                # print('new_items: ', new_items)
                for item in new_items:
                    new_qualifier = '&' + item
                    new_items[new_items.index(item)] = new_qualifier

                if len(new_items) > 1:
                    qualifier_list.pop(i)
                    qualifier_list = qualifier_list + new_items
                    # qualifier_list[i:i + 1] = new_items

        # print('qualifier_list', qualifier_list)

    return qualifier_list


#
# Get Qualiders to be replaced
#
def get_qualifier_overrides(qualifier_list, JOB_NAME, PROC_NAME, check_jcl):

    # print('Inside get_qualifier_overrides', qualifier_list, JOB_NAME, PROC_NAME)
    prcpath =  r"C:\Users\UX169817\PyCharm Project\Automation\tar\20231205_vg-project-all-code"
    os.chdir(prcpath)
    if check_jcl is False:
        proc_name = PROC_NAME + '.PRC'

    else:
        proc_name = JOB_NAME + '.JCL'
    # proc_name = PROC_NAME + '.PRC'
    ACTUAL_QUALIFIER = ''
    ACTUAL_QUALIFIER_LIST = []
    PARMfoundinJCL = False

    for qualifier in qualifier_list:
        actual_qualifier_found = False
        prefix = ''
        temp_qualifier = ''
        
    # to handle alias which starts in between a word like 'V&PRD'
    if qualifier[0] == '&':
        temp_qualifier = qualifier[1:]
        # temp_qualifier = re.sub('\W+', '', qualifier)
        # print('temp_qualifier ', temp_qualifier)

    else:
        pos = qualifier.find('&')

        # check if the '&' symbol is presnt in the word
        if pos >= 0:
            # extract the part of the word before the '&' symbol
            prefix = qualifier[:pos]
            temp_qualifier = qualifier[pos + 1:]
        # print('temp_qulifier: ', temp_qualifier)
        ACTUAL_QUALIFIER, PARMfoundinJCL = check_jcl_override(JOB_NAME, temp_qualifier, PROC_NAME)

        if PARMfoundinJCL is True:
            if qualifier[0] != '&':
                ACTUAL_QUALIFIER = prefix + ACTUAL_QUALIFIER

            # print('ACTUAL_QUALIFIER: ', ACTUAL_QUALIFIER)
            ACTUAL_QUALIFIER_LIST.append(ACTUAL_QUALIFIER)
            actual_qualifier_found = True

        if PARMfoundinJCL is False:
            with open(f"{prcpath}\{proc_name}", encoding='utf-8', errors='ignore') as linesproc:
                temp_qualifier = temp_qualifier + '='
                # print('temp_qualifier PROC check', temp_qualifier)
                for lineproc in linesproc:
                    # print(lineproc)
                    if lineproc.find(temp_qualifier) != -1 and lineproc.find('=') != -1 and lineproc[0:3] != "//*":
                       split_lineproc = lineproc.split(',')
                       # print('lineproc', lineproc)
                       # print(len(split_lineproc), split_lineproc)
                       if len(split_lineproc) > 0:
                           for each_split in split_lineproc:
                               # print('each_split: ', each_split)
                               if each_split.find(temp_qualifier) != -1 and lineproc[0:3] != "//*":
                                   if each_split.find("'") != -1:
                                      # print("found ''' ")
                                      # print(each_split.count("'"))
                                      if each_split.find("'") != 1:
                                          # MODULE_NAME = each_split[each_split.find("'") + 1:each_split.find(",") - 1]
                                          match = re.search(r"'(.*?)'", each_split)
                                          ACTUAL_QUALIFIER = match.group(1) if match else  ''
                                          # print(MODULE_NAME)
                                          if qualifier[0] != '&':
                                              ACTUAL_QUALIFIER = prefix + ACTUAL_QUALIFIER
                                              # print('job name PARM: ', JOB_NAME)
                                          # print('PROC Name PARM: ', PROC_NAME)
                                          #     print('temp_module Name PARM :', temp_qualifier)
                                          # print('Module Name PARM: ', ACTUAL_QUALIFIER)
                                          ACTUAL_QUALIFIER_LIST.append(ACTUAL_QUALIFIER)
                                          actual_qualifier_found = True
                                          break
                                   else:
                                       # print('else:')
                                       match = re.search(r'(.+)=([^=]*)', each_split)

                                       if match:
                                           # extract the second string part
                                           ACTUAL_QUALIFIER = match.group(2)

                                           # print the second string part
                                           # print('actual_qualifier_found: ', ACTUAL_QUALIFIER)
                                           if all(char.isspace() for char in ACTUAL_QUALIFIER) and len(
                                                   ACTUAL_QUALIFIER) > 0:
                                               pass
                                           else:
                                               if '  ' in ACTUAL_QUALIFIER:
                                                   split_second_string = ACTUAL_QUALIFIER.split()
                                                   ACTUAL_QUALIFIER = split_second_string[0]
                                                   # print(ACTUAL_QUALIFIER)

                                               if ',' in ACTUAL_QUALIFIER:
                                                   ACTUAL_QUALIFIER = ACTUAL_QUALIFIER[:-1]
                                                   if ',' in ACTUAL_QUALIFIER:
                                                       ACTUAL_QUALIFIER = ACTUAL_QUALIFIER[:-1]
                                                   # print('remove ",":v ', ACTUAL_QUALIFIER)
                                               if '\n' in ACTUAL_QUALIFIER:
                                                   ACTUAL_QUALIFIER = ACTUAL_QUALIFIER[:-1]

                                               # ACTUAL_QUALIFIER = each_split[each_split.find("=") + 1:each_split.find(",")]
                                               # if len(ACTUAL_QUALIFIER) > 0:
                                               #     ACTUAL_QUALIFIER_TEMP = ACTUAL_QUALIFIER.split()
                                               #     ACTUAL_QUALIFIER = ACTUAL_QUALIFIER_TEMP[0]
                                               # ACTUAL_QUALIFIER = re.sub('\W+', '', ACTUAL_QUALIFIER)
                                               # ACTUAL_QUALIFIER = ACTUAL_QUALIFIER[0:8]
                                               if qualifier[0] != '&':
                                                   ACTUAL_QUALIFIER = prefix + ACTUAL_QUALIFIER
                                               # print('Job name PARM: ', JOB_NAME)
                                               # print('PROC name PARM: ', PROC_NAME)
                                               # print('temp_module Name PARM : ', temp_qualifier)
                                               # print('MODULE Name PARM1 : ', ACTUAL_QUALIFIER)
                                               ACTUAL_QUALIFIER_LIST.append(ACTUAL_QUALIFIER)
                                               actual_qualifier_found = True
                                               break
                    if actual_qualifier_found is True:
                        break

        if actual_qualifier_found is False:
            if temp_qualifier == 'SYSUID':
                ACTUAL_QUALIFIER_LIST.append(temp_qualifier)
            else:
                 # print('else qualifeir: ', qualifier)
                 # print('else qualifier list', qualifier_list)
                 ACTUAL_QUALIFIER_LIST.append(qualifier)
                 # qualifier_list.remove(qualifier)

    return ACTUAL_QUALIFIER_LIST

###***********************************
### Get input files from PROC
###***********************************
#
def extract_files(Job_NAME, PROC_NAME, check_jcl, proc_step_name):
    prcpath = r"C:\Users\UX169817\PyCharm Project\Automation\tar\20231205_vg-project-all-code"
    os.chdir(prcpath)
    global job_location
    if check_jcl is False:
        proc_name = PROC_NAME + '.PRC'

    else:
        proc_name = Job_NAME + '.JCL'

    jcl_step = proc_step_name
    srch_txt_DDDSN = 'DSN='
    srch_txt_DDDISP = 'DISP=SHR'
    srcht_txt_DISPOLD = 'DISP=OLD'
    srch_txt_amphnd = '(&'
    srch_txt_exec = ' EXEC '
    line_num = 0
    utility_name = ''
    INPUT_FILE_LIST = []
    ACTUAL_INPUT_FILE_LIST = []
    INPUT_STEP_LIST = []
    logical_file_name_list = []
    disposition_parameter = ''
    disposition_list = []
    access_type_list = []
    qualifier_list = []
    ACTUAL_QUALIFIER_LIST = []
    actual_split_file_list = []
    ACTUAL_INPUT_FILE_LIST  = []
    ACTUAL_UTILITY_LIST = []
    FORMATTED_FILE_LIST = []
    global next_line
    next_line = 0
    current_step_name = ''
    previous_line_file = ''
    logical_file_name = ''
    prev_logical_file_name = ''
    ACTUAL_INPUT_FILE_LIST_override = []
    actual_qualifier_found = False
    file_in_next_line = False
    prev_line_minus_two = ''
    prev_line_minus_one = ''
    prev_line = ''
    dd_blocks = []
    exec_blocks = []
    current_block = []

    print('proc_name: ', proc_name)
    try:
        with open(f"{prcpath}\{proc_name}", encoding='utf-8', errors='ignore') as file:
            lines = file.readlines()
            # Extract Each steps of the PROC as blocks
            l1 = 0
            for line in lines:
                # print('line: ', line)
                if ' EXEC ' in line and line[0:3] != "// " and line[0:3] != "//*":

                    if current_block:
                        # if 'EXEC' in current_block:
                        exec_blocks.append(current_block)
                        current_block = []
                        # current_block = []
                    current_block.append(line.strip())
                elif current_block:
                    if line[0:3] == "// ":
                        # print(line)
                        characters = line[2:70] # Extract characters from the 3rd to the 72nd position
                        if characters.isspace():
                            lines_missed = len(lines) - l1
                            # print(lines_missed)
                            if lines_missed > 4:
                                JOB_COMMENT = '//Null line found in JCL/PROC'
                                write_output_file(application, schedule, lob, Job_NAME, PROC_NAME. '', '', '', '',
                                                    '', JOB_COMMENT, '', '')
                                JOB_COMMENT = ''
                            break
                        current_block.append(line.strip())
                    l1 = l1 + 1

            # Add the last block if it exists
            if current_block:
                exec_blocks.append(current_block)

            # exec_blocks will have all the steps of the PROC.
            current_block = []

            # Go through each steps of the PROC and extract DD blocks (where the file is)
            # print(len(exec_blocks))
            if check_jcl is True:
                if len(exec_blocks) == 0:
                    JOB_COMMENT = "No executable step found in the job"
                    num_of_row_written = write_output_file(application, schedule, lob, Job_NAME, PROC_NAME, '',
                                                           '', '', '', utility_name, JOB_COMMENT, current_step_name, '')

        for blocks in exec_blocks:
            # print('******************************** Step Block Starts **********************************')
            dd_blocks = []
            current_block = []
            utility_name = ''
            step_file_count = 0
            for line in block:
                # print(line)
                # form each step of the PROC, extract the utility being executed
                if line.find(srch_txt_exec) != -1:
                    if line[0:3] != "//*" and line[0:3] != "// ":
                        split_line = line.split()
                        current_step_name = split_line[0]
                        current_step_name = re.sub('\w+', '', current_step_name)
                        # print('current_step_name', current_step_name)
                        utility_name = line[line.find(" EXEC ") + 6:]
                        # print('utility_name1: ', utility_name)
                        if utility_name[0] == ' ':
                            utility_name = utility_name[1:]

                        if ',' in utility_name:
                            split_utility = utility_name.split(',')
                            utility_name = split_utility[0]

                        if check_jcl is True:
                            if 'PROC=' in utility_name:
                                utility_name = utility_name[utility_name.find('=') + 1:]

                        if ' ' in utility_name:
                            utility_name = utilty_name.split(' ')[0]

                        utility_name = utility_name.strip()

                        # print('utility_name: ', utility_name)
                if ' DD ' in line and line[0:3] != '//*':
                    if current_block:
                        # if 'EXEC in current_block:
                        dd_blocks.append(current_block)
                        current_block = []
                        # current_block = []
                    current_block.append(line.strip())
                elif current_block:
                    current_block.append(line.strip())

                # Add the last block if it exists
            if current_block:
                dd_blocks.append(current_block)
                # print(line)

            # dd_blocks

            for dd_blocks in dd_blocks:
                disposition_parameter = ''
                logical_file_name = ''
                access_type = ''
                temp_file_name = ''
                formatted_temp_file = ''
                # prev_logical_file_name = ''
                # print('********* dd block starts ****************')
                # print('current_step_name', current_step_name)
                for lineproc in dd_blocks:
                    # print(lineproc)

                    # Get logical file name
                    if  lineproc[0:3] != "//*" and lineproc[0:3] != "// " and lineproc[0:3] != "/* " and lineproc.find(' DD ') != -1:
                        logical_file_name = lineproc[lineproc.find("//") + 2:lineproc.find(" DD")]
                        logical_file_name_list = logical_file_name.strip()
                        prev_logical_file_name = logical_file_name

                        # print('logical_file_name: ', logical_file_name)

                    # Get Dispositions
                    if lineproc[0:3] != "//*" and lineproc.find('DISP=') != -1:
                        # print(lineproc)
                        disposition_parameter = lineproc[lineproc.find("DISP=") + 0;]


                        if lineproc.find('DISP=(') != -1:
                            if lineproc.find('DISP=(OLD,DELETE,KEEP)') != -1 or lineproc.find('DISp=(OLD,DELETE') != -1 or lineproc.find('DISP=(SHR') != -1 or lineproc.find('DISP=(OLD,KEEP') != -1 or lineproc.find('DISP=(OLD,PASS' != -1 or lineproc.find('DISP=(OLD)') != -1):
                                access_type = 'Read'
                            else:
                                access_type = 'Write'

                            disposition_parameter = disposition_parameter.split(' ')[0]
                            disposition_parameter = disposition_parameter.split(')')[0]

                        elif lineproc.find(srch_txt_DDDISP) != -1 or lineproc.find(srcht_txt_DISPOLD) != -1:
                            access_type = 'Read'
                            disposition_parameter = disposition_parameter.split(',')[0]
                            disposition_parameter = disposition_parameter.split(' ')[0]

                        elif lineproc.find('DISP=MOD'):
                            access_type = 'Write'
                            disposition_parameter = disposition_parameter.split(',')[0]
                            disposition_parameter = disposition_parameter.split(' ')[0]

                        # print('disposition_parameter : ', disposition_parameter)

                        # check if 'DSN=' present on a current line and extract the file
                        if lineproc.find(srch_txt_DDDSN) != -1:
                            if lineproc[0:3] != "//*":
                                # print('lineprcoc: ',lineproc)
                                temp_file_name = lineproc[lineproc.find("DSN=") + 4:lineproc.find(",")]
                                temp_file_name = lineproc[lineproc.find("DSN=") + 4]
                                # temp_file_name = tmep_file_name
                                # print('temp_file_name: ', temp_file_name)
                                if len(temp_file_name) == 0:
                                    temp_file_name = lineproc[lineproc.find("DSN=") + 4:]
                                    if ',' in temp_file_name:
                                        split_file = temp_file_name.split(',')
                                        temp_file_name = split_file[0]

                                if ',' in temp_file_name:
                                    split_file =temp_file_name.split(',')
                                    temp_file_name = split_file[0]

                                if ' 'in temp_file_name:
                                    temp_file_name = temp_file_name.split(' ')[0]

                                if temp_file_name.find('(') != -1:
                                    temp_file_name = lineproc[lineproc.find("DSN=") + 4:lineproc.find("(")]
                                    temp_file_name = re.sub(',','', temp_file_name)
                                    # temp_file_name = temp_file_name[:-4]
                                # print('File extract ', temp_file_name)

                                # if temp_file_name not in INPUT_FILE_LIST:
                                # if temp_file_name.find('PARM') != -1 or temp_file_name.find('SYS4.') != -1 or temp_file_name.find('CARDLIB') != -1 or logical_file_name == 'JOBLIB' or temp_file_name.find('SMTF.DATA') != -1:
                                #   pass
                                # else:
                                if '\n' in temp_file_name:
                                    temp_file_name = temp_file_name[:-1]

                                split_temp_file_name = temp_file_name.split('.')
                                if len(split_temp_file_name) == 1:
                                    temp_file_name = check_set_from_jcl(split_temp_file_name[0], JOB_NAME)

                                formatted_temp_file = handle_aphrnd_period(temp_file_name)
                                if '\n' in formatted_temp_file:
                                    formatted_temp_file = formatted_temp_file[:-1]

                                # print('temp_file_name: ', temp_file_name, formatted_temp_file)

                    if len(logical_file_name) == 0:
                        logical_file_name = prev_logical_file_name

                    if check_jcl is True:
                        if len(utility_name) > 3 and utility_name[3] =='P' and utility_name not in [ 'MISPSMTP', 'IEBPTPPCH', 'DB2CPY1P']:
                            pass
                        else:
                            if len(temp_file_name) > 0:
                                # print('temp_file_name ', temp_file_name)
                                # print('logical_file_name: ', logical_file_name)
                                # print('disposition_parameter: ', disposition_parameter)
                                # print('current_step_name: ', current_step_name)
                                # print('utility_name: ', utility_name)
                                INPUT_FILE_LIST.append(temp_file_name)
                                FORMATTED_FILE_LIST.append(current_step_name)
                                INPUT_STEP_LIST.append(current_step_name)
                                ACTUAL_UTILITY_LIST.append(utility_name)
                                logical_file_name_list.append(logical_file_name)
                                disposition_list.append(disposition_parameter)
                                access_type_list.append(access_type)
                                step_file_count = step_file_count + 1
                                # print(temp_file_name, utility_name)

                    else:
                        if len(temp_file_name) > 0:
                                # print('temp_file_name ', temp_file_name)
                                # print('logical_file_name: ', logical_file_name)
                                # print('disposition_parameter: ', disposition_parameter)
                                # print('access_type ', access_type)
                                # print('utility_name: ', utility_name)
                                INPUT_FILE_LIST.append(temp_file_name)
                                FORMATTED_FILE_LIST.append(current_step_name)
                                INPUT_STEP_LIST.append(current_step_name)
                                ACTUAL_UTILITY_LIST.append(utility_name)
                                logical_file_name_list.append(logical_file_name)
                                disposition_list.append(disposition_parameter)
                                access_type_list.append(access_type)
                                step_file_count = step_file_count + 1

                    g# print('************************ dd block Ends ****************************')
                # print('****************************** step block ends *************************')

                # even if there is no files used in the step, write the step along with utility to the o/p
                # print('step_file_count: ',step_file_count)
                if step_file_count == 0:
                    if check_jcl is True:
                        # print('utility_name: ',utility_name, logical_file_name)
                        if len(utility_name) > 3 and utility_name[3] == 'P' and utility_name not in ['MISPSMTP', 'IEBPTPCH', 'DB2CPY1P'] or '.' in logical_file_name_list########
                            pass
                        else:
                            # print('no File Found in JCL step')
                            JOB_COMMENT = 'No filesfound in the step'
                            num_of_row_written = write_output_file(application, schedule, lob, JOB_NAME, PROC_NAME,'' 
                                                                   '','','',utility_name,JOB_COMMENT,current_step_name,'')
                    else:
                        JOB_COMMENT = 'No files found in the step'
                        num_of_row_written = write_output_file(application, schedule, lob, Job_NAME, PROC_NAME, '',
                                                               current_step_name, '', '', utility_name, JOB_COMMENT,
                                                               jcl_step, '')

                #print('INPUT_FILE_LIST: ', INPUT_FILE_LIST)
            # print(len(INPUT_FILE_LIST))

            # check if there are any file ovverrides from JCL
            qualifier_override = False
            INPUT_FILE_LIST, FORMATTED_FILE_LIST = check_jcl_full_file_override(Job_NAME, PROC_NAME, INPUT_FILE_LIST,
                                                                                INPUT_STEP_LIST, logical_file_name_list,
                                                                                FORMATTED_FILE_LIST, qualifier_override)
            # INPUT_FILE_LIST = INPUT_FILE_LIST_override
            # print('INPUT_FILE_LIST: ', INPUT_FILE_LIST)
            # from each file extracted, get the qualifiers to be replaced
            qualifier_list = get_qualifiers(INPUT_FILE_LIST)

            if len(qualifier_list) > 0:
                sorted_qualifier_list = sorted(qualifier_list, key=lambda x: len(x), reverse=True)
                qualifier_list = sorted_qualifier_list

            # print('sprted_qualifier_list: ', qualifier_list)

            ACTUAL_QUALIFIER_LIST = get_qualifier_overrides(qualifier_list, Job_NAME, PROC_NAME, check_jcl)

            # print('ACTUAL_QUALIFIER_LIST: ',ACTUAL_QUALIFIER_LIST)

            actual_temp_qual_list = []
            for qual in ACTUAL_QUALIFIER_LIST:
                # print('qual: ', qual)
                temp_qual_list = []
                actual_qual = qual
                if '&' in qual:
                    formatted_temp_qual = handle_aphrnd_period(qual)
                    # print('formatted_temp_qual: ', formatted_temp_qual)
                    temp_qual_list.append(qual)
                    qual_list_temp = get_qualifiers(temp_qual_list)
                    actual_qual_list_temp = get_qualifier_override(qual_list_temp, Job_NAME, PROC_NAME, check_jcl)
                    # print('actual_qual_list_temp: ', actual_qual_list_temp)

                    actual_qual = formatted_temp_qual
                    for i in range(len(qual_list_temp)):
                        actual_qual = actual_qual.replace(qual_list_temp[i], actual_qual_list_temp[i])

                # print('actual_qual: ',actual_qual)
                actual_temp_qual_list.append(actual_qual)

            ACTUAL_QUALIFIER_LIST = actual_temp_qual_list
            for file in FORMATTED_FILE_LIST:
                ACTUAL_INPUT_FILE = file
                for i in range(len(qualifier_list)):
                    # print(i)
                    # print('qualifier: ', qualifier_list)
                    # print('actual_qualifier: ', ACTUAL_QUALIFIER_LIST)
                    # print('ACTUAL_OUTPUT_FILE: ', ACTUAL_INPUT_FILE)
                    # print('qualifier_list[i]: ',qualifier_list[i] )
                    # print('ACTUAL_QUALIFIER_LIST[i]: ', ACTUAL_QUALIFIER_LIST[i])
                    ACTUAL_INPUT_FILE_LIST = ACTUAL_INPUT_FILE.replace(qualifier_list[i], ACTUAL_QUALIFIER_LIST[i])
                    # print('ACTUAL_OUTPUT_FILE: ', ACTUAL_INPUT_FILE)
                    # print('qualifier_list[i]: ', qualifier_list[i])
                    # print('ACTUAL_QUALIFIER_LIST[i]: ', ACTUAL_QUALIFIER_LIST[i])
                # print('length of file : ', len(ACTUAL_INPUT_FILE))
                if ACTUAL_INPUT_FILE.find('(') != -1:
                    ACTUAL_INPUT_FILE = ACTUAL_INPUT_FILE[:ACTUAL_INPUT_FILE.find("(")]

                ACTUAL_INPUT_FILE_LIST.append(ACTUAL_INPUT_FILE)
                # print('ACTUAL_QUALIFIER_LIST: ', ACTUAL_QUALIFIER_LIST)
                # print('INPUT_FILE_LIST: ', INPUT_FILE_LIST)
                # print('ACTUAL_INPUT_FILE_LIST: ', ACTUAL_INPUT_FILE_LIST)
                qualifier_override = True
                ACTUAL_INPUT_FILE_LIST_override, FORMATTED_FILE_LIST = check_jcl_full_file_override(Job_NAME, PROC_NAME,
                                                                                                    ACTUAL_INPUT_FILE_LIST,
                                                                                                    INPUT_STEP_LIST,
                                                                                                    logical_file_name_list,
                                                                                                    FORMATTED_FILE_LIST,
                                                                                                    qualifier_override)
                ACTUAL_INPUT_FILE_LIST = ACTUAL_INPUT_FILE_LIST_override

                # print('qualifier_list: ', qualifier_list)
                # print('ACTUAL_QUALIFIER_LIST: ', ACTUAL_QUALIFIER_LIST)
                # print('INPUT_FILE_LIST: ', INPUT_FILE_LIST)
                # print('ACTUAL_INPUT_FILE_LISTL: ',ACTUAL_INPUT_FILE_LIST)
                # print("ACTUAL_INPUT_FILE_LIST_override: ", ACTUAL_INPUT_FILE_LIST_override)

                return ACTUAL_INPUT_FILE_LIST, INPUT_STEP_LIST, logical_file_name_list, ACTUAL_UTILITY_LIST, disposition_list, access_type_list
                # return ACTUAL_INPUT_FILE_LIST, INPUT_STEP_LIST, logical_file_name_list, ACTUAL_UTILITY_LIST

    except FileNotFoundError:
        # print(' coming here')
        # PROC_NAME not in ('IDCAMS', 'IEFBR14', 'IEBGENER', 'SORT', 'FLUSH', 'SAS'] and
        if check_jcl is True:
            JOB_COMMENT = 'Job not found'
            rows_written = write_output_file(application, schedule, Job_NAME, '', '', '' , '', '', '', JOB_COMMENT,
                            '', '')

        else:
            if len(PROC_NAME) > 3 and PROC_NAME[3] == 'P' and PROC_NAME[0:3] !='EZT':
                print('PROC Not Found')
                JOB_COMMENT = 'PROC not found'
                rows_written = write_output_file(application, schedule, lob, Job_NAME, PROC_NAME, '', '', '', '', '', JOB_COMMENT, '', '')
                JOB_COMMENT = ''
            return ACTUAL_INPUT_FILE_LIST, INPUT_STEP_LIST, logical_file_name_list, ACTUAL_UTILITY_LIST, disposition_list, access_type_list
def handle_aphrnd_period(temp_file_name):
    remove_first_period = False
    temp_file = ''
    for i in range(len(temp_file_name)):
        # print(text[i])
        letter = temp_file_name[i]

        if temp_file_name[i] == '&':
            remove_first_period = True
        if temp_file_name[i] == '.' and remove_first_period is True:
            letter = ''
            remove_first_period = False

        temp_file = temp_file + letter
        # print(temp_file)
    return temp_file


# ##***********************************************************
# ##                            Main flow
# ##***********************************************************
pathXX =os.path.abspath(os.getcwd())
num_of_row_written = 0
UTILITY_NAME = ''
IKJEFT_PGM = ''
JOB_COMMENT = ''
qualifier_list = []
check_jcl = False
global check_jcl_flag
check_jcl_flag = True
global application, schedule, lob, isn, sub_application, parent_folder, data_center
application = ''
schedule = ''
lob = ''
isn = ''
sub_application = ''
parent_folder = ''
data_center = ''

try:
    c=0
    for row in input_sheet.iter_rows(min_row=1, min_col=1, max_row=mrow, max_col=mcol):
        c = c + 1
        colindex = 1
        # for cell in row:
        # global application, schedule, lob
        check_jcl_flag = True
        isn = row[0].value
        application = row[1].value
        schedule = row[2].value
        tmp_jobname = row[3].value
        parent_folder = row[4].value
        data_center = row[5].value
        # schedule = row[1].value
        lob = row[6].value
        job_location = ''
        print(f"input Job_name - {c} :", tmp_jobname)
        # print('input app :', application)
        # print('input schedule :', schedule)
        # print('input lob :', lob)
        # MAIN_JOBNAME, MAIN_PROC_LIST, proc_step_list = get_proc_name(tmp_jobname)
        get_proc_name(tmp_jobname)


finally:

    os.chdir(pathdir)
    # result_wb.save("JobFileList2.xlsx")
    # result_wb.close()
    print('Please wait, loading the pandas data to excel......!!!')
    if len(records) > 1000000:
        midpoint = len(records) - 1000000

        records1 = records[:midpoint]
        records2 = records[midpoint:]

        df_output = pd.DataFrame(records1)
        df_output.to_excel("JobFilesList2_Part1.xlsx", index=False)

        df_output = pd.DataFrame(records2)
        df_output.to_excel("JobFilesList2_Part2.xlsx", index=False)
    else:

        df_output = pd.DataFrame(records)
        df_output.to_excel("JobFilesList2.xlsx", index=False)

    print('Done')


    end_time = time.time()

    # Calculate the elapsed time
    elapsed_time = end_time - start_time
    elapsed_minutes = elapsed_time / 60

    print('num_of_row_in_list : ', len(records))
    print(f"Script took {elapsed_minutes:.2f} minutes to complete.")








































